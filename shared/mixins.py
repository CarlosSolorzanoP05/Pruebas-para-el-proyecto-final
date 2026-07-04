from django.contrib import messages
from django.shortcuts import redirect
from django.http import HttpResponse
import io
from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import redirect
from django.contrib import messages

# ─────────────────────────────────────────────
#  StaffRequiredMixin  (sin cambios)
# ─────────────────────────────────────────────
class StaffRequiredMixin:
    """
    Mixin que verifica si el usuario es miembro del staff.
    Si no es staff, redirige con mensaje de error.
    """
    staff_redirect_url = '/'
    staff_error_message = 'You do not have permission to perform this action. Staff access required.'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_staff:
            messages.error(request, self.staff_error_message)
            return redirect(self.staff_redirect_url)
        return super().dispatch(request, *args, **kwargs)


# ─────────────────────────────────────────────
#  Helpers de exportación genéricos (PDF / Excel)
#  Usados tanto por ExportMixin (CBV) como por las
#  vistas basadas en función (FBV) a través de
#  export_list_response().
# ─────────────────────────────────────────────

def resolve_field(obj, field):
    """Obtiene el valor de un campo simple, relación (__) o callable."""
    if callable(field):
        return field(obj)
    value = obj
    for part in field.split('__'):
        value = getattr(value, part, '')
        if callable(value):
            value = value()
    return value if value is not None else ''


def build_rows(qs, fields):
    """Construye las filas de datos (lista de listas de strings)."""
    rows = []
    for obj in qs:
        row = [str(resolve_field(obj, f)) for f in fields]
        rows.append(row)
    return rows


def build_excel_response(qs, filename, headers, fields):
    """Genera un HttpResponse .xlsx a partir de un queryset."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse(
            'openpyxl no está instalado. Ejecuta: pip install openpyxl',
            status=500,
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename[:31]   # max 31 chars en Excel

    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(color='FFFFFF', bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, row in enumerate(build_rows(qs, fields), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response


def build_pdf_response(qs, filename, headers, fields):
    """Genera un HttpResponse .pdf con el listado tabulado."""
    try:
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
        )
        from reportlab.lib.styles import getSampleStyleSheet
    except ImportError:
        return HttpResponse(
            'reportlab no está instalado. Ejecuta: pip install reportlab',
            status=500,
        )

    rows = build_rows(qs, fields)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(filename.upper(), styles['Title']))
    elements.append(Spacer(1, 0.4 * cm))

    table_data = [headers] + rows
    num_cols = len(headers)
    col_width = (landscape(letter)[0] - 3 * cm) / max(num_cols, 1)

    table = Table(table_data, colWidths=[col_width] * num_cols, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, 0), 9),
        ('ALIGN',      (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME',   (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',   (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EBF5FB')]),
        ('GRID',       (0, 0), (-1, -1), 0.4, colors.HexColor('#AED6F1')),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING',    (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    return response


def export_list_response(request, qs, filename, headers, fields):
    """
    Punto de entrada genérico para FBVs: revisa ?export=pdf|excel y
    devuelve el HttpResponse correspondiente, o None si no aplica
    (para que la vista siga su flujo normal de render).

    Uso típico en una vista de función:
        def my_list(request):
            qs = MyModel.objects.all()
            export = export_list_response(
                request, qs, 'listado_x', ['Col1', 'Col2'], ['field1', 'field2']
            )
            if export:
                return export
            return render(request, 'app/my_list.html', {'items': qs})
    """
    fmt = request.GET.get('export', '').lower()
    if fmt == 'pdf':
        return build_pdf_response(qs, filename, headers, fields)
    if fmt == 'excel':
        return build_excel_response(qs, filename, headers, fields)
    return None


# ─────────────────────────────────────────────
#  ExportMixin  —  genérico para cualquier ListView
# ─────────────────────────────────────────────
class ExportMixin:
    """
    Mixin genérico para exportar el queryset filtrado a PDF o Excel.

    Uso en una ListView:
        class MiListView(LoginRequiredMixin, ExportMixin, ListView):
            export_filename = 'mi_listado'   # sin extensión
            export_headers  = ['Col1', 'Col2', ...]
            export_fields   = ['campo1', 'campo2', ...]
            # export_fields puede contener:
            #   - nombre de atributo simple          → 'name'
            #   - lookups con doble guión bajo        → 'brand__name'
            #   - callables que reciben el objeto     → lambda obj: ...

    El mixin intercepta ?export=pdf o ?export=excel en la URL.
    Si no hay ese parámetro, sigue el flujo normal de ListView.
    """

    export_filename = 'listado'
    export_headers: list = []
    export_fields: list = []

    # ── punto de entrada ──────────────────────────────────────────────
    def get(self, request, *args, **kwargs):
        fmt = request.GET.get('export', '').lower()
        if fmt in ('pdf', 'excel'):
            qs = self._get_export_queryset(request)
            if fmt == 'pdf':
                return build_pdf_response(qs, self.export_filename, self.export_headers, self.export_fields)
            return build_excel_response(qs, self.export_filename, self.export_headers, self.export_fields)
        return super().get(request, *args, **kwargs)

    # ── extrae queryset filtrado (reutiliza get_queryset() de la vista) ─
    def _get_export_queryset(self, request):
        # get_queryset() ya aplica los filtros definidos en la vista
        return self.get_queryset()
class GroupRequiredMixin(LoginRequiredMixin):
    """
    Mixin que permite el acceso solo a usuarios que pertenecen a ciertos grupos.
    Uso en la vista: group_required = ['Administrador', 'Vendedor']
    """
    group_required = None

    def dispatch(self, request, *args, **kwargs):
        # 1. Verificar si está logueado (heredado de LoginRequiredMixin)
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        
        # 2. Si el usuario es superusuario (admin total), pasa directo
        if request.user.is_superuser:
            return super().dispatch(request, *args, **kwargs)

        # 3. Validar si se definieron grupos requeridos
        if self.group_required:
            # Obtenemos los nombres de los grupos del usuario actual
            user_groups = request.user.groups.values_list('name', flat=True)
            
            # Verificamos si hay coincidencia entre los grupos requeridos y los del usuario
            if not any(group in user_groups for group in self.group_required):
                messages.error(request, "No tienes permisos (Rol adecuado) para acceder a esta pantalla.")
                return redirect('home') # O la url de tu página de inicio/dashboard
                
        return super().dispatch(request, *args, **kwargs)